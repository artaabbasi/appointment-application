import inspect
import io

from openpyxl.workbook import Workbook

from common.excel_styles import header_style, primary_style, secondary_style


class ExcelExportUtil:
    def __init__(self):
        self.header_style = header_style
        self.primary_style = primary_style
        self.secondary_style = secondary_style

    async def get_list_export(self, column_data: list, objs) -> io.BytesIO:
        """Export data to Excel with proper handling of both synchronous and asynchronous functions."""
        workbook = Workbook()
        sheet = workbook.worksheets[0]

        # Write headers
        sheet.cell(row=1, column=1, value="#").style = self.header_style
        for col_num, col_name, _ in column_data:
            sheet.cell(row=1, column=col_num, value=col_name).style = self.header_style

        # Process and write data rows
        for row_idx, obj in enumerate(objs, start=2):
            row_style = self.primary_style if row_idx % 2 != 0 else self.secondary_style

            # Write row number
            sheet.cell(row=row_idx, column=1, value=f"{row_idx - 1}").style = row_style

            # Process each column
            for col_num, _, func in column_data:
                cell = sheet.cell(row=row_idx, column=col_num)
                cell.style = row_style

                try:
                    # Get the raw function result first
                    result = func(obj)

                    # Handle async cases
                    if inspect.iscoroutine(result):
                        cell.value = await result
                    elif inspect.iscoroutinefunction(func):
                        cell.value = await func(obj)
                    else:
                        cell.value = result

                except Exception as e:
                    cell.value = f"Error: {str(e)}"
                    # Consider logging the error here as well

        # Auto-adjust column widths
        for column_cells in sheet.columns:
            if not column_cells:
                continue

            column = column_cells[0].column_letter
            max_length = max(
                (len(str(cell.value)) for cell in column_cells if cell.value is not None),
                default=0
            )
            sheet.column_dimensions[column].width = (max_length + 2) * 1.2

        # Final worksheet configuration
        sheet.freeze_panes = "A2"

        # Save to BytesIO
        output_stream = io.BytesIO()
        workbook.save(output_stream)
        output_stream.seek(0)

        return output_stream