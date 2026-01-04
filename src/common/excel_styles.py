from openpyxl.styles import NamedStyle
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill


header_style = NamedStyle(name="header_style")
header_style.alignment = Alignment(horizontal="center", vertical="center")
header_style.font = Font(bold=False, size=14, color="000000")
header_style.fill = PatternFill("solid", start_color="E1F5F2")
header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )


primary_style = NamedStyle(name="primary_style")
primary_style.alignment = Alignment(horizontal="center", vertical="center")
primary_style.font = Font(bold=False, size=14, color="000000")
primary_style.fill = PatternFill("solid", start_color="FAFAFA")
primary_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )


secondary_style = NamedStyle(name="secondary_style")
secondary_style.alignment = Alignment(horizontal="center", vertical="center")
secondary_style.font = Font(bold=False, size=14, color="000000")
secondary_style.fill = PatternFill("solid", start_color="FFFFFF")
secondary_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )